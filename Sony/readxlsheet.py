import openpyxl
from openpyxl.reader.excel import load_workbook

#file="C:/Users/deepa/PycharmProjects/Myphthon/Sony/datasheet.xlsx"
file="C:/Users/deepa/PycharmProjects/Myphthon/Sony/datasheet1.xlsx"
workbook=openpyxl.load_workbook(file)
sheet=workbook.active
#same date into excel sheet
# for r in range(1,80):
#     for c in range(1,8):
#         sheet.cell(r,c).value="Prakash"
# workbook.save(file)
# workbook.close()

#updated Excel sheet one by one data
sheet.cell(1,1).value=123
sheet.cell(1,2).value="Johan"
sheet.cell(1,3).value="Civil-Engineer"


sheet.cell(2,1).value=143
sheet.cell(2,2).value="Smith"
sheet.cell(2,3).value="Hard-Engineer"


sheet.cell(3,1).value=144
sheet.cell(3,2).value="Cavin"
sheet.cell(3,3).value="Soft-Engineer"

sheet.cell(4,1).value=154
sheet.cell(4,2).value="Rajkumar"
sheet.cell(4,3).value="Soft-Engineer"


workbook.save(file)



