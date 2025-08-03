#Reading Excel
#Write Excel
import openpyxl
import csv

# set the path for xl sheet.
file="C:/Users/deepa/PycharmProjects/Myphthon/Sony/datasheet.xlsx"
workbook=openpyxl.load_workbook(file)
sheet=workbook.active
rows = sheet.max_row
cols = sheet.max_column


for r in range(1,rows+1):
    for c in range(1,cols+1):
        print(sheet.cell(r,c).value,end="                          ")
    print()


